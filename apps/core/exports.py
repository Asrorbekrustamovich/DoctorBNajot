"""Word va Excel eksport qilish uchun yordamchi funksiyalar."""
import base64
import re
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


# data:image/png;base64,.... ko'rinishidagi rasmlar (imzo shu tarzda saqlanadi)
_DATA_IMG_RE = re.compile(
    r'src\s*=\s*(["\'])\s*data:image/(?P<ext>png|jpeg|jpg|gif)\s*;\s*base64\s*,(?P<data>[^"\']+)\1',
    re.IGNORECASE,
)


def _extract_data_images(html: str):
    """HTML ichidagi base64 rasmlarni ajratib oladi.

    Word (MS Word) `data:` URI'ni umuman ko'rsatmaydi — shuning uchun rasmlarni
    alohida MHTML qismlariga chiqaramiz va HTML'da oddiy fayl nomiga almashtiramiz.
    Natija: (yangi_html, [(fayl_nomi, mime, base64_data), ...])
    """
    parts = []

    def _repl(m):
        ext = m.group("ext").lower()
        if ext == "jpg":
            ext = "jpeg"
        data = re.sub(r"\s+", "", m.group("data"))
        name = f"image{len(parts) + 1:03d}.{'jpg' if ext == 'jpeg' else ext}"
        parts.append((name, f"image/{ext}", data))
        return f'src="{name}"'

    return _DATA_IMG_RE.sub(_repl, html), parts


def _b64_lines(data: str) -> str:
    """Base64 ma'lumotni 76 belgidan qatorlarga bo'ladi (MIME talabi)."""
    return "\n".join(data[i:i + 76] for i in range(0, len(data), 76))


def export_html_to_word(html_content: str, filename: str) -> HttpResponse:
    """HTML kontentni MS Word tushunadigan .doc formatida qaytaradi.

    Agar HTML ichida base64 rasm (imzo) bo'lsa — hujjat MHTML
    (multipart/related) ko'rinishida yig'iladi, shunda imzo rasmi Word'da
    ekrandagidek ko'rinadi.
    """
    html_content, images = _extract_data_images(html_content)

    word_html = f"""<html xmlns:o='urn:schemas-microsoft-com:office:office'
xmlns:w='urn:schemas-microsoft-com:office:word'
xmlns='http://www.w3.org/TR/REC-html40'>
<head>
<meta charset="utf-8">
<style>
    body {{ font-family: 'Times New Roman', serif; font-size: 14pt; color: #000; }}
    table {{ border-collapse: collapse; width: 100%; margin-bottom: 15px; }}
    th, td {{ border: 1px solid #000; padding: 5px; }}
    h1, h2, h3, h4 {{ color: #000; text-align: center; }}
    /* Rasmiy hujjatda tugma/menyu/interaktiv narsalar KO'RINMASIN */
    .d-print-none, .btn, button, .modal, .dbn-sidebar, nav, .navbar, form button {{ display: none !important; }}
    @page Section1 {{ size: 8.5in 11.0in; margin: 1.0in 1.0in 1.0in 1.0in; mso-header-margin: .5in; mso-footer-margin: .5in; mso-paper-source: 0; }}
    div.Section1 {{ page: Section1; }}
</style>
</head>
<body>
<div class="Section1">
{html_content}
</div>
</body>
</html>"""

    if not images:
        response = HttpResponse(word_html, content_type='application/msword; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.doc"'
        return response

    # --- MHTML (Word o'zi shu formatda saqlaydi) ---
    boundary = "----=_NextPart_EduMedHIS"
    base = "file:///C:/edumed/doc"
    html_b64 = _b64_lines(base64.b64encode(word_html.encode("utf-8")).decode("ascii"))

    chunks = [
        "MIME-Version: 1.0",
        f'Content-Type: multipart/related; boundary="{boundary}"; type="text/html"',
        "X-Document-Type: Word.Document",
        "",
        f"--{boundary}",
        f"Content-Location: {base}/main.htm",
        'Content-Type: text/html; charset="utf-8"',
        "Content-Transfer-Encoding: base64",
        "",
        html_b64,
        "",
    ]
    for name, mime, data in images:
        chunks += [
            f"--{boundary}",
            f"Content-Location: {base}/{name}",
            f"Content-Type: {mime}",
            "Content-Transfer-Encoding: base64",
            "",
            _b64_lines(data),
            "",
        ]
    chunks.append(f"--{boundary}--")

    mhtml = "\r\n".join("\r\n".join(c.split("\n")) for c in chunks)
    response = HttpResponse(mhtml, content_type='application/msword')
    response['Content-Disposition'] = f'attachment; filename="{filename}.doc"'
    return response


def export_queryset_to_excel(queryset, columns: list, filename: str) -> HttpResponse:
    """
    QuerySet ni Excel faylga o'girib beradi.
    columns: [(Header, Field/Property name), ...]
    Masalan: [("F.I.SH", "patient.full_name"), ("Sana", "visit_date")]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    # Headerlar
    header_font = Font(bold=True)
    for col_num, (header_title, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ma'lumotlar
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, field_path) in enumerate(columns, 1):
            val = obj
            try:
                for part in field_path.split('.'):
                    if val is None:
                        break
                    # dict bo'lsa
                    if isinstance(val, dict):
                        val = val.get(part)
                    else:
                        val = getattr(val, part)
                    if callable(val):
                        val = val()
            except AttributeError:
                val = ""
            
            # Sana formatini to'g'rilash
            if hasattr(val, "strftime"):
                if hasattr(val, "hour"):
                    val = val.strftime("%d.%m.%Y %H:%M")
                else:
                    val = val.strftime("%d.%m.%Y")
                
            ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else "")

    # Ustun kengligini to'g'rilash
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
