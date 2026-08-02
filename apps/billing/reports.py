"""Direktor uchun hisobotlar: tushum statistikasi, shifokor narxlari, Excel.

Tushum ikki ko'rinishda hisoblanadi:
  - Hisoblangan (accrued): cheklardagi umumiy summa
  - To'langan (kassa):     amalda qabul qilingan pul

Shifokor kesimidagi tushum Consultation.fee snapshotlaridan olinadi —
narx keyin o'zgarsa ham tarixiy hisobotlar o'zgarmaydi.
"""
from __future__ import annotations

import datetime
from decimal import Decimal

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, DecimalField, F, Q, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.generic import TemplateView

from apps.accounts.models import Role, User
from apps.accounts.permissions import RoleRequiredMixin, role_required
from apps.billing.models import Invoice, InvoiceItem
from apps.clinical.models import Consultation, DoctorPrice, DoctorPriceHistory
from apps.patients.models import Patient
from apps.registration.models import Visit

REPORT_ROLES = (Role.Code.DIRECTOR, Role.Code.SUPER_ADMIN, Role.Code.ACCOUNTANT)
PRICE_ROLES = (Role.Code.DIRECTOR, Role.Code.SUPER_ADMIN, Role.Code.ADMINISTRATOR)

_DEC = DecimalField(max_digits=14, decimal_places=2)


def _period_range(request) -> tuple[datetime.date, datetime.date, str]:
    """?period=week|month|day yoki ?start=&end= dan sana oralig'ini oladi."""
    today = timezone.localdate()
    period = request.GET.get("period", "month")
    start_s, end_s = request.GET.get("start"), request.GET.get("end")

    if start_s and end_s:
        try:
            start = datetime.date.fromisoformat(start_s)
            end = datetime.date.fromisoformat(end_s)
            if start > end:
                start, end = end, start
            return start, end, "custom"
        except ValueError:
            pass  # noto'g'ri sana — default davrga tushamiz

    if period == "day":
        return today, today, "day"
    if period == "week":
        start = today - datetime.timedelta(days=today.weekday())
        return start, today, "week"
    if period == "year":
        return today.replace(month=1, day=1), today, "year"
    # default: joriy oy
    return today.replace(day=1), today, "month"


def build_report(start: datetime.date, end: datetime.date) -> dict:
    """Berilgan oraliq uchun barcha hisobot ma'lumotlari."""
    invoices = Invoice.objects.filter(
        created_at__date__gte=start, created_at__date__lte=end
    ).exclude(status=Invoice.Status.CANCELLED)

    totals = invoices.aggregate(
        accrued=Coalesce(Sum("total_amount"), Value(Decimal(0), output_field=_DEC)),
        paid=Coalesce(Sum("paid_amount"), Value(Decimal(0), output_field=_DEC)),
        refunded=Coalesce(Sum("refunded_amount"), Value(Decimal(0), output_field=_DEC)),
    )
    # Sof tushum = to'langan - qaytarilgan
    totals["net_paid"] = totals["paid"] - totals["refunded"]
    totals["debt"] = totals["accrued"] - totals["net_paid"]

    # Kunlik seriya (diagramma uchun)
    daily = list(
        invoices.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            accrued=Coalesce(Sum("total_amount"), Value(Decimal(0), output_field=_DEC)),
            paid=Coalesce(Sum("paid_amount"), Value(Decimal(0), output_field=_DEC)),
            refunded=Coalesce(Sum("refunded_amount"), Value(Decimal(0), output_field=_DEC)),
        )
        .order_by("day")
    )
    by_day = {d["day"]: d for d in daily}
    labels, paid_series, accrued_series = [], [], []
    d = start
    while d <= end:
        labels.append(d.strftime("%d.%m"))
        row = by_day.get(d)
        # Kunlik sof tushum (qaytarishlar ayirilgan)
        paid_series.append(float(row["paid"] - row["refunded"]) if row else 0.0)
        accrued_series.append(float(row["accrued"]) if row else 0.0)
        d += datetime.timedelta(days=1)

    # Shifokorlar kesimida (fee snapshotlardan — tarixiy narxlar o'zgarmaydi)
    doctors = list(
        Consultation.objects.filter(
            created_at__date__gte=start, created_at__date__lte=end
        )
        .values("doctor__id", "doctor__first_name", "doctor__last_name")
        .annotate(
            visits=Count("id"),
            income=Coalesce(Sum("fee"), Value(Decimal(0), output_field=_DEC)),
        )
        .order_by("-income")
    )

    # Jarrohlik daromadi (jarroh kesimida) — shifokor ulushiga qo'shiladi
    from apps.clinical.models import SurgerySchedule
    surgery_by_doctor = {
        row["surgeon__id"]: row
        for row in SurgerySchedule.objects.filter(
            created_at__date__gte=start, created_at__date__lte=end
        ).exclude(status="cancelled")
        .values("surgeon__id", "surgeon__first_name", "surgeon__last_name")
        .annotate(
            surgeries=Count("id"),
            surgery_income=Coalesce(Sum("actual_price"), Value(Decimal(0), output_field=_DEC)),
        )
    }

    THIRD = Decimal("3")
    seen_ids = set()
    for row in doctors:
        row["name"] = f"{row['doctor__last_name']} {row['doctor__first_name']}".strip() or "—"
        s = surgery_by_doctor.get(row["doctor__id"])
        row["surgeries"] = s["surgeries"] if s else 0
        row["surgery_income"] = s["surgery_income"] if s else Decimal(0)
        row["total_income"] = row["income"] + row["surgery_income"]
        # SHIFOKOR ULUSHI: (qabul + operatsiya) summasining 1/3 qismi
        row["doctor_share"] = (row["total_income"] / THIRD).quantize(Decimal("0.01"))
        row["clinic_share"] = row["total_income"] - row["doctor_share"]
        seen_ids.add(row["doctor__id"])

    # Faqat operatsiya qilgan (ambulator qabuli bo'lmagan) jarrohlar ham kirsin
    for doc_id, s in surgery_by_doctor.items():
        if doc_id in seen_ids:
            continue
        total = s["surgery_income"]
        doctors.append({
            "doctor__id": doc_id,
            "name": f"{s['surgeon__last_name']} {s['surgeon__first_name']}".strip() or "—",
            "visits": 0,
            "income": Decimal(0),
            "surgeries": s["surgeries"],
            "surgery_income": total,
            "total_income": total,
            "doctor_share": (total / THIRD).quantize(Decimal("0.01")),
            "clinic_share": total - (total / THIRD).quantize(Decimal("0.01")),
        })
    doctors.sort(key=lambda r: r["total_income"], reverse=True)

    # YOTISH (statsionar) va UMUMIY DAVOLANISH tushumini ALOHIDA hisoblash
    inpatient_total = InvoiceItem.objects.filter(
        invoice__in=invoices, item_type=InvoiceItem.ItemType.INPATIENT
    ).aggregate(t=Coalesce(Sum("total_price"), Value(Decimal(0), output_field=_DEC)))["t"]
    treatment_total = InvoiceItem.objects.filter(
        invoice__in=invoices
    ).exclude(item_type=InvoiceItem.ItemType.INPATIENT).aggregate(
        t=Coalesce(Sum("total_price"), Value(Decimal(0), output_field=_DEC))
    )["t"]

    # Chek bandlari turlari bo'yicha (Xizmat/Dori/Statsionar)
    by_type = list(
        InvoiceItem.objects.filter(invoice__in=invoices)
        .values("item_type")
        .annotate(total=Coalesce(Sum("total_price"), Value(Decimal(0), output_field=_DEC)))
        .order_by("-total")
    )
    type_labels = dict(InvoiceItem.ItemType.choices)
    for row in by_type:
        row["label"] = type_labels.get(row["item_type"], row["item_type"])

    # Eng ko'p buyurilgan xizmatlar (top 10)
    top_services = list(
        InvoiceItem.objects.filter(invoice__in=invoices)
        .values("name")
        .annotate(
            cnt=Count("id"),
            total=Coalesce(Sum("total_price"), Value(Decimal(0), output_field=_DEC)),
        )
        .order_by("-total")[:10]
    )

    # Tashrif statistikasi
    visits_qs = Visit.objects.filter(visit_date__gte=start, visit_date__lte=end)
    visit_stats = {
        "total": visits_qs.count(),
        "completed": visits_qs.filter(status=Visit.Status.COMPLETED).count(),
        "cancelled": visits_qs.filter(status=Visit.Status.CANCELLED).count(),
        "open": visits_qs.filter(status__in=(
            Visit.Status.CREATED, Visit.Status.WAITING,
            Visit.Status.ACCEPTED, Visit.Status.IN_PROGRESS,
        )).count(),
    }
    new_patients = Patient.objects.filter(
        created_at__date__gte=start, created_at__date__lte=end
    ).count()

    return {
        "start": start,
        "end": end,
        "totals": totals,
        "chart_labels": labels,
        "chart_paid": paid_series,
        "chart_accrued": accrued_series,
        "doctors": doctors,
        "by_type": by_type,
        "top_services": top_services,
        "visit_stats": visit_stats,
        "new_patients": new_patients,
        "invoice_count": invoices.count(),
        # Yotish (statsionar) va davolanish tushumi ALOHIDA
        "inpatient_total": inpatient_total,
        "treatment_total": treatment_total,
    }


class RevenueReportView(RoleRequiredMixin, TemplateView):
    """Direktor hisobot paneli: tushum, diagrammalar, statistika."""

    allowed_roles = REPORT_ROLES
    template_name = "billing/revenue_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start, end, period = _period_range(self.request)
        context.update(build_report(start, end))
        context["period"] = period
        return context


@role_required(*REPORT_ROLES)
def revenue_excel(request):
    """Hisobotni Excel (.xlsx) faylga eksport qiladi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    start, end, _ = _period_range(request)
    data = build_report(start, end)

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2E7D32")
    title_font = Font(bold=True, size=13)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")

    # 1-varaq: Umumiy
    ws = wb.active
    ws.title = "Umumiy"
    ws["A1"] = f"Tushum hisoboti: {start:%d.%m.%Y} — {end:%d.%m.%Y}"
    ws["A1"].font = title_font
    rows = [
        ("Hisoblangan tushum (cheklar)", float(data["totals"]["accrued"])),
        ("To'langan (kassa)", float(data["totals"]["paid"])),
        ("Qaytarilgan (refund)", float(data["totals"]["refunded"])),
        ("Sof tushum (to'langan - qaytarilgan)", float(data["totals"]["net_paid"])),
        ("Qarzdorlik", float(data["totals"]["debt"])),
        ("— Faqat yotishdan (statsionar) tushum", float(data["inpatient_total"])),
        ("— Umumiy davolanishdan tushum (yotishsiz)", float(data["treatment_total"])),
        ("Cheklar soni", data["invoice_count"]),
        ("Tashriflar (jami)", data["visit_stats"]["total"]),
        ("Yakunlangan tashriflar", data["visit_stats"]["completed"]),
        ("Bekor qilingan tashriflar", data["visit_stats"]["cancelled"]),
        ("Ochiq tashriflar", data["visit_stats"]["open"]),
        ("Yangi bemorlar", data["new_patients"]),
    ]
    ws.append([]); ws.append(["Ko'rsatkich", "Qiymat"])
    style_header(ws, 3, 2)
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18

    # 2-varaq: Kunlik tushum
    ws2 = wb.create_sheet("Kunlik")
    ws2.append(["Sana", "Hisoblangan", "To'langan"])
    style_header(ws2, 1, 3)
    for lbl, acc, paid in zip(data["chart_labels"], data["chart_accrued"], data["chart_paid"]):
        ws2.append([lbl, acc, paid])
    for i in range(1, 4):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    # 3-varaq: Shifokorlar kesimida (1/3 ulush bilan)
    ws3 = wb.create_sheet("Shifokorlar")
    ws3.append([
        "Shifokor", "Qabullar", "Qabul tushumi", "Operatsiyalar",
        "Operatsiya tushumi", "Jami", "Shifokor ulushi (1/3)", "Klinika (2/3)",
    ])
    style_header(ws3, 1, 8)
    for row in data["doctors"]:
        ws3.append([
            row["name"], row["visits"], float(row["income"]),
            row.get("surgeries", 0), float(row.get("surgery_income", 0)),
            float(row.get("total_income", row["income"])),
            float(row.get("doctor_share", 0)), float(row.get("clinic_share", 0)),
        ])
    ws3.column_dimensions["A"].width = 28
    for col in "BCDEFGH":
        ws3.column_dimensions[col].width = 17

    # 4-varaq: Xizmatlar
    ws4 = wb.create_sheet("Xizmatlar")
    ws4.append(["Nomi", "Soni", "Jami summa"])
    style_header(ws4, 1, 3)
    for row in data["top_services"]:
        ws4.append([row["name"], row["cnt"], float(row["total"])])
    ws4.column_dimensions["A"].width = 44
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="tushum_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"'
    )
    wb.save(response)
    return response


@role_required(*PRICE_ROLES)
def doctor_prices(request):
    """Shifokorlar qabul narxlarini kiritish/yangilash.

    Narx o'zgartirilganda avvalgi qabullar (fee snapshot) o'zgarmaydi,
    o'zgarish DoctorPriceHistory ga yoziladi.
    """
    if request.method == "POST":
        doctor_id = request.POST.get("doctor_id")
        price_str = request.POST.get("price", "").replace(",", ".").replace(" ", "").strip()
        doctor = get_object_or_404(
            User, id=doctor_id, is_active=True,
            role__code__in=(Role.Code.DOCTOR, Role.Code.CHIEF_DOCTOR, Role.Code.SURGEON),
        )
        try:
            price = Decimal(price_str)
        except Exception:
            messages.error(request, "Narxni to'g'ri kiriting.")
            return redirect("billing:doctor_prices")
        if price < 0:
            messages.error(request, "Narx manfiy bo'lishi mumkin emas.")
            return redirect("billing:doctor_prices")

        with transaction.atomic():
            dp, created = DoctorPrice.objects.select_for_update().get_or_create(
                doctor=doctor, defaults={"price": price}
            )
            old_price = Decimal(0) if created else dp.price
            if not created and dp.price != price:
                dp.price = price
                dp.is_active = True
                dp.save(update_fields=["price", "is_active"])
            if created or old_price != price:
                DoctorPriceHistory.objects.create(
                    doctor=doctor, old_price=old_price,
                    new_price=price, changed_by=request.user,
                )
        messages.success(
            request,
            f"{doctor.get_full_name()} qabul narxi: {price} so'm "
            "(avvalgi qabullarga ta'sir qilmaydi).",
        )
        return redirect("billing:doctor_prices")

    doctors = (
        User.objects.filter(
            is_active=True,
            role__code__in=(Role.Code.DOCTOR, Role.Code.CHIEF_DOCTOR, Role.Code.SURGEON),
        )
        .select_related("role", "consultation_price")
        .order_by("last_name", "first_name")
    )
    history = (
        DoctorPriceHistory.objects.select_related("doctor", "changed_by")
        .order_by("-created_at")[:30]
    )
    return render(request, "billing/doctor_prices.html", {
        "doctors": doctors,
        "history": history,
    })
